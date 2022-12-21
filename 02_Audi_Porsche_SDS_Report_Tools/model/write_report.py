#! /usr/bin/env python
# -*- encoding:utf-8 -*-
from typing import Any

import openpyxl
# from Get_Config import Config
from model.Get_Config import Config

config = Config()

# summary: Any = {'word': 5, 'sentence': 8, 'nlu_stage': 12, 'nlu_live': 14}
summary: Any = {'his_word': 5, 'word': 6, 'cont_word': 8, 'his_sentence': 10, 'sentence': 11, 'cont_sentence': 13,
                'his_nlu_stage': 16, 'nlu_stage': 17, 'cont_nlu_stage': 19, 'his_nlu_live': 20, 'nlu_live': 21,
                'cont_nlu_live': 23}

index_testsuit = 1
index_category = 2
data_area_start = 19
if config.project.lower().strip() in ['audi', 'audiclu37']:
    data_area_end = 31
elif config.project.lower().strip() in ['porsche']:
    data_area_end = 23


class WriteResult(object):
    def __init__(self, dir_result_report):
        self.dir_result_report = dir_result_report
        self.wb = openpyxl.load_workbook(dir_result_report)
        # ws = self.wb.active
        self.sheets = self.wb.sheetnames
        self.sum_ws = self.wb[self.sheets[0]]

    def write_summary(self, testsuit, envoriment, sum_data, hisresult):
        # envoriment = stage
        rowinfo = self.get_row_bycategory(self.get_row_bysuit(testsuit.lower()))
        # print(rowinfo)  # [('null', 19), ('nlu', 20), ('nlu_speech', 21), ('总计', 22)]
        # print(sum_data)  # {'nlu': [['nlu', '92.94%'], ['nlu_speech', '94.66%'], ['总计', '94.33%']]}
        for k, v in sum_data.items():  # item()函数读取sum_data字典，返回单个元组如('nlu'，[['nlu', '92.94%'], ['nlu_speech', '94.66%'], ['总计', '94.33%']])
            for item in rowinfo:  # 遍历rowinfo列表
                for data in v:  # 遍历上面的元组第二个参数，也就是列表
                    if item[0] == data[0]:  # 如果rowinfo列表 == 元素列表;rowinfo['null'] == ['nlu', '92.94%']中的第一个元素
                        if envoriment != None:  # 并且传参过来的envoriment不是none,一般是stage
                            # 则在xls第零sheet的row行cloumn列以cell方法写入value=data[1]；如row=19，column=nlu_stage
                            if hisresult == 0:
                                self.sum_ws.cell(row=item[1], column=summary[k + '_' + envoriment]).value = data[1]
                            else:
                                self.sum_ws.cell(row=item[1], column=summary['his_' + k + '_' + envoriment]).value = \
                                data[1]
                                # his_nlu_stage
                        else:
                            if hisresult == 0:
                                self.sum_ws.cell(row=item[1], column=summary[k]).value = data[1]
                            else:
                                self.sum_ws.cell(row=item[1], column=summary['his_' + k]).value = data[1]


    def write_data(self, langauge: object, enviroment: object, testtype: object, datas: object) -> object:
        sheet_name = self.get_sheetname(langauge, enviroment, testtype)
        print(sheet_name)
        ws = self.wb[sheet_name]
        # 是否已经有数据，有数据去掉第一行再写
        if ws.max_row > 1:
            datas = datas[1:]
        start_row = ws.max_row - 1
        for i, data in enumerate(datas):
            for j, d in enumerate(data):
                ws.cell(row=start_row + 1 + i, column=j + 1).value = str(d)


    # NEW
    def count_difference(self, vlu1, vlu2):
        # 执行计算，返回计算结果
        vlue1, vlue2 = float(vlu1.strip('%')) * 100, float(vlu2.strip('%')) * 100
        count = vlue2 - vlue1
        if count == 0:
            return "0.00%"
        else:
            count = str(count / 100) + "%"
            return count


    # NEW
    def write_count_difference(self, countype):
        # print("write_count_difference")
        try:
            if countype == 'word':
                self.get_sum_result(summary['his_word'], summary['word'], summary['cont_word'])
            elif countype == 'sentence':
                self.get_sum_result(summary['his_sentence'], summary['sentence'], summary['cont_sentence'])
            elif countype == 'nlu_stage':
                self.get_sum_result(summary['his_nlu_stage'], summary['nlu_stage'], summary['cont_nlu_stage'])
            elif countype == 'nlu_live':
                self.get_sum_result(summary['his_nlu_live'], summary['nlu_live'], summary['cont_nlu_live'])
            else:
                print('get_sum_result error')
        except:
            print('write_count_difference error!')

    # NEW
    def get_sum_result(self, his, cur, cou):
        try:
            for i in range(19, 31):
                # print("\n开始算差值：")
                print(i, his, cur, cou)
                val1 = str(self.sum_ws.cell(row=i, column=his).value)
                val2 = str(self.sum_ws.cell(row=i, column=cur).value)
                if val1 == '/' or val2 == '/':
                    # print("填写 / ")
                    self.sum_ws.cell(row=i, column=cou).value = '/'
                elif val1 and val2:
                    count = self.count_difference(val1, val2)
                    print(val1, val2, count)
                    self.sum_ws.cell(row=i, column=cou).value = count
                else:
                    # print("填入 null ")
                    self.sum_ws.cell(row=i, column=cou).value = 'null'
                # print("算差值完毕！\n")
        except:
            print('get_sum_result error!')

    def get_row_bysuit(self, testsuit):
        # 该方法用于定位传参testsuit在xls中的行数
        # 读取第0个sheet，从data_area_start=19，遍历到data_area_end=Audi=31，获取列表，row=19~31=12，index_testsuit=1
        # 也就是说获取xls中第一列的说19到31的值.
        testsuits = [self.sum_ws.cell(row=row, column=index_testsuit).value for row in
                     range(data_area_start, data_area_end)]
        # print(testsuits) # ['Mandarin', None, None, None, 'Cantonese', None, None, None, '氛围灯_中文', '氛围灯_粤语',
        cols = {}
        start = data_area_start
        temp = []
        # 遍历第一列的值，如果是none则使用上一个元素的值，此处是用于处理单元格合并的问题。
        for item in testsuits:
            if item == None:
                cols[temp[-1]].append(start)  # cols = {Mandarin：20}
            else:
                # xx.setdefault(aa,bb),字典setdefault()方法和get()方法类似, 如果键不存在于字典中，将会添加键并将值设为默认值.
                # 设置字典cols键为item值为[]，并apped(start)到前面的[] # {'mandarin': [19]} # {'mandarin': [19, 20]}
                cols.setdefault(item.lower(), []).append(start)
                # 只要不是None，temp列表添加item元素，相当于获取了绝对值。
                temp.append(item.lower())
            # 循环一次 + 1，目的是为了让cols字典中列表值对应xls的每个位置
            start += 1
        # 最后返回cols字典中传参过来testsuit对应的值。如testsuit=氛围灯训练集_中文，则返回29
        # print("get_row_bysuit == ",cols[testsuit])
        return cols[testsuit]

    def get_row_bycategory(self, rows):
        # 假如rows为29行，常量index_category = 2，也就是说获取rows行的第二列元素和row行数
        category_list = [(self.sum_ws.cell(row=row, column=index_category).value, row) for row in rows]
        # print('get_row_bycategory↓↓↓')
        # print (category_list)
        return category_list

    def get_sheetname(self, language, environment, testtype):
        for sheet in self.sheets:
            if testtype == 'asr' and '%s_%s' % (language, testtype) == sheet:
                return sheet
            elif testtype == 'nlu' and '%s_%s_%s' % (language, testtype, environment) == sheet:
                return sheet

    def save(self, dir_report):
        self.wb.save(dir_report)


if __name__ == '__main__':
    FILE = r'/home/xslan/Documents/AUDI&Porsche_SDS_Report_Tools/Result/Audi_CW15_SDS自动化测试报告.xlsx'
    wr = WriteResult(FILE)
    # result = [['null', '98.43%'], ['nlu', '99.52%'], ['nlu_speech', '98.95%'], ['综合', '98.75%'], ['', '']]
    print(wr.count_difference('92.55%','92.55%'))